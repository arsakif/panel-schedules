"""
Module for writing extracted panel schedule data to Excel files.
Format: Concatenated header in first column, circuits below, 2 empty rows between panels.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Any
import re


def clean_ocp_size(ocp_value: str) -> str:
    """
    Extract just the numeric value from OCP size.
    Examples: "20A" -> "20", "60.00A" -> "60", "30 Amps" -> "30"
    """
    if not ocp_value:
        return ""
    # Extract number from string
    match = re.search(r'(\d+(?:\.\d+)?)', ocp_value)
    if match:
        number = float(match.group(1))
        # Remove .00 decimals if present
        if number == int(number):
            return str(int(number))
        else:
            return str(number)
    return ocp_value  # Return original if no number found


class ExcelWriter:
    """Handles writing panel schedule data to Excel files."""
    
    def __init__(self, filename: str = "panel_schedules.xlsx"):
        """
        Initialize Excel writer.
        
        Args:
            filename: Output filename for the Excel file
        """
        self.filename = filename
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Panel Schedules"
        self.current_row = 1
        
        # Define styles
        self.default_font = Font(name='Arial', size=11)
        self.bold_font = Font(name='Arial', size=11, bold=True)
        self.panel_info_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        self.circuit_header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def write_panel(self, panel: Dict[str, Any]) -> None:
        """
        Write a complete panel to Excel matching combined CSV structure:
        - Row 1: [Empty] [Empty] [1] [Panel header info concatenated]
        - Row 2: [Empty] [Panel Name] [2] [Load Description] [Quantity] [OCP] [Poles] [Feeder] [Ckt#]
        - Row 3+: [Empty] [Panel name] [Empty] [Load desc] [1] [OCP] [Poles] [Feeder] [Ckt#]
        - 2 empty rows after each panel
        
        Args:
            panel: Dictionary containing panel header and circuits
        """
        panel_header = panel.get("panel_header", {})
        circuits = panel.get("circuits", [])
        panel_name = panel_header.get("panel_name", "")
        
        # Build concatenated panel header text (shortened format)
        header_parts = []
        if panel_name:
            header_parts.append(f"Pnl: {panel_name}")
        if panel_header.get("bus_amperage"):
            header_parts.append(f"Bus/Lug: {panel_header['bus_amperage']}")
        if panel_header.get("main_ocpd"):
            # Extract just amperage value from main_ocpd
            main_ocpd = panel_header['main_ocpd']
            match = re.search(r'(\d+(?:\.\d+)?)\s*A', main_ocpd)
            if match:
                amperage = float(match.group(1))
                if amperage == int(amperage):
                    main_ocpd = f"{int(amperage)}A"
                else:
                    main_ocpd = f"{amperage}A"
            header_parts.append(f"Main: {main_ocpd}")
        if panel_header.get("voltage"):
            header_parts.append(panel_header['voltage'])
        if panel_header.get("phase"):
            header_parts.append(panel_header['phase'])
        if panel_header.get("wire"):
            header_parts.append(panel_header['wire'])
        if panel_header.get("poles"):
            header_parts.append(f"Poles: {panel_header['poles']}")
        if panel_header.get("kaic"):
            header_parts.append(panel_header['kaic'])
        if panel_header.get("enclosure"):
            header_parts.append(panel_header['enclosure'])
        
        panel_description = " | ".join(header_parts)
        
        # Track panel start row for border application
        panel_start_row = self.current_row
        
        # Row 1: Panel header with structure [Empty] [Empty] [1] [Panel description]
        # Apply green fill (92D050) to columns 2-9
        for col_idx in range(1, 10):
            cell = self.sheet.cell(row=self.current_row, column=col_idx)
            cell.font = self.default_font
            if col_idx >= 2:  # Borders only on columns 2-9
                cell.border = self.border
            if col_idx == 3:
                cell.value = "1"
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 4:
                cell.value = panel_description
                cell.font = self.bold_font
                cell.alignment = Alignment(wrap_text=True)
            if col_idx >= 2:
                cell.fill = self.panel_info_fill
        self.current_row += 1
        
        # Row 2: Circuit headers - Apply yellow fill (FFFF00) to columns 2-9
        circuit_headers = ["", "Panel Name", "2", "Load Description", "Quantity", "OCP", "Poles", "Feeder", "Ckt#"]
        for col_idx in range(1, 10):
            cell = self.sheet.cell(row=self.current_row, column=col_idx)
            if col_idx >= 2:  # Borders only on columns 2-9
                cell.border = self.border
            if col_idx <= len(circuit_headers):
                cell.value = circuit_headers[col_idx - 1]
                if circuit_headers[col_idx - 1]:
                    cell.font = self.bold_font
                else:
                    cell.font = self.default_font
            # Center align columns 3, 5, 6, 7, 9
            if col_idx in [3, 5, 6, 7, 9]:
                cell.alignment = Alignment(horizontal='center')
            if col_idx >= 2:
                cell.fill = self.circuit_header_fill
        self.current_row += 1
        
        # Write each circuit
        if circuits:
            for circuit in circuits:
                # Apply borders and font to all cells in row (columns 1-9)
                for col_idx in range(1, 10):
                    cell = self.sheet.cell(row=self.current_row, column=col_idx)
                    cell.font = self.default_font
                    if col_idx >= 2:
                        cell.border = self.border
                    # Center align columns 3, 5, 6, 7, 9
                    if col_idx in [3, 5, 6, 7, 9]:
                        cell.alignment = Alignment(horizontal='center')
                
                self.sheet.cell(row=self.current_row, column=1, value="")
                self.sheet.cell(row=self.current_row, column=2, value=panel_name)
                self.sheet.cell(row=self.current_row, column=3, value="")
                self.sheet.cell(row=self.current_row, column=4, value=circuit.get("load_description", ""))
                # Quantity as number
                self.sheet.cell(row=self.current_row, column=5, value=1)
                # OCP as number
                ocp_value = clean_ocp_size(circuit.get("ocp_size", ""))
                try:
                    self.sheet.cell(row=self.current_row, column=6, value=int(ocp_value) if ocp_value else "")
                except (ValueError, TypeError):
                    self.sheet.cell(row=self.current_row, column=6, value=ocp_value)
                # Poles as number
                poles_value = circuit.get("poles", "")
                try:
                    self.sheet.cell(row=self.current_row, column=7, value=int(poles_value) if poles_value else "")
                except (ValueError, TypeError):
                    self.sheet.cell(row=self.current_row, column=7, value=poles_value)
                self.sheet.cell(row=self.current_row, column=8, value=circuit.get("feeder", ""))
                self.sheet.cell(row=self.current_row, column=9, value=circuit.get("circuit_number", ""))
                self.current_row += 1
        else:
            # No circuits found
            for col_idx in range(1, 10):
                cell = self.sheet.cell(row=self.current_row, column=col_idx)
                cell.font = self.default_font
                if col_idx >= 2:
                    cell.border = self.border
                # Center align columns 3, 5, 6, 7, 9
                if col_idx in [3, 5, 6, 7, 9]:
                    cell.alignment = Alignment(horizontal='center')
            self.sheet.cell(row=self.current_row, column=1, value="")
            self.sheet.cell(row=self.current_row, column=2, value=panel_name)
            self.sheet.cell(row=self.current_row, column=3, value="")
            self.sheet.cell(row=self.current_row, column=4, value="NO CIRCUITS FOUND")
            self.current_row += 1
        
        # Add 2 empty rows between panels
        self.current_row += 2
        """
        Write a complete panel (header + circuits) to Excel.
        
        Args:
            panel: Dictionary containing panel header and circuits
        """
        # Write panel header
        panel_header = panel.get("panel_header", {})
        # Add 2 empty rows between panels
        self.current_row += 2
    
    def write_all_panels(self, panels: List[Dict[str, Any]]) -> None:
        """
        Write all panels to the Excel file.
        
        Args:
            panels: List of panel dictionaries
        """
        for panel in panels:
            self.write_panel(panel)
    
    def adjust_column_widths(self) -> None:
        """Auto-adjust column widths for better readability."""
        column_widths = {
            1: 8.33,   # 1st column
            2: 16.85,  # 2nd column
            3: 8.33,   # 3rd column
            4: 65,     # 4th column
            5: 8,    # 5th column
            6: 8,    # 6th column
            7: 8,    # 7th column
            8: 25,     # 8th column
            9: 7.5     # 9th column
        }
        
        for col_idx, width in column_widths.items():
            column_letter = self.sheet.cell(row=1, column=col_idx).column_letter
            self.sheet.column_dimensions[column_letter].width = width
    
    def save(self) -> None:
        """Save the Excel workbook to file."""
        self.adjust_column_widths()
        
        # Turn off gridlines for better appearance
        self.sheet.sheet_view.showGridLines = False
        
        self.workbook.save(self.filename)
        print(f"Excel file saved: {self.filename}")


